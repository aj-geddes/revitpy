"""Output formatters for different file formats."""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Union

import pandas as pd

from .models import RoomData


class CSVFormatter:
    """Format room data as CSV."""

    def write(
        self,
        rooms: List[RoomData],
        output_path: Union[str, Path],
        columns: Optional[List[str]] = None,
    ) -> None:
        """Write room data to CSV file.

        Args:
            rooms: List of room data
            output_path: Path to output file
            columns: List of columns to include
        """
        if not rooms:
            return

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to dictionaries
        room_dicts = [room.to_dict(columns) for room in rooms]

        # Use pandas for clean CSV writing
        df = pd.DataFrame(room_dicts)

        # Reorder columns if specified
        if columns:
            available_columns = [col for col in columns if col in df.columns]
            df = df[available_columns]

        df.to_csv(output_path, index=False)


class ExcelFormatter:
    """Format room data as Excel with formatting."""

    def write(
        self,
        rooms: List[RoomData],
        output_path: Union[str, Path],
        include_formatting: bool = True,
        include_summary: bool = True,
    ) -> None:
        """Write room data to Excel file.

        Args:
            rooms: List of room data
            output_path: Path to output file
            include_formatting: Apply formatting
            include_summary: Include summary sheet
        """
        if not rooms:
            return

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to DataFrame
        room_dicts = [room.to_dict() for room in rooms]
        df = pd.DataFrame(room_dicts)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Write main data sheet
            df.to_excel(writer, sheet_name="Rooms", index=False)

            # Get workbook and worksheet for formatting
            if include_formatting:
                workbook = writer.book
                worksheet = writer.sheets["Rooms"]

                # Apply formatting
                self._apply_formatting(worksheet, df)

            # Add summary sheet if requested
            if include_summary:
                summary_df = self._create_summary(df)
                summary_df.to_excel(writer, sheet_name="Summary")

    def _apply_formatting(self, worksheet, df):
        """Apply Excel formatting.

        Args:
            worksheet: Excel worksheet
            df: DataFrame
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-size columns
        for col_num, column_title in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                df[column_title].astype(str).apply(len).max(),
                len(column_title)
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Format numeric columns
        for col_num, column_title in enumerate(df.columns, 1):
            if column_title in ["Area", "Volume"]:
                for row_num in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.number_format = "0.00"

    def _create_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Create summary statistics.

        Args:
            df: Room data DataFrame

        Returns:
            Summary DataFrame
        """
        summary_data = {
            "Metric": [
                "Total Rooms",
                "Total Area",
                "Average Area",
                "Total Volume",
                "Average Volume",
            ],
            "Value": [
                len(df),
                df["Area"].sum() if "Area" in df.columns else 0,
                df["Area"].mean() if "Area" in df.columns else 0,
                df["Volume"].sum() if "Volume" in df.columns else 0,
                df["Volume"].mean() if "Volume" in df.columns else 0,
            ],
        }

        return pd.DataFrame(summary_data)


class JSONFormatter:
    """Format room data as JSON."""

    def write(
        self,
        rooms: List[RoomData],
        output_path: Union[str, Path],
        include_metadata: bool = True,
        pretty_print: bool = True,
    ) -> None:
        """Write room data to JSON file.

        Args:
            rooms: List of room data
            output_path: Path to output file
            include_metadata: Include metadata
            pretty_print: Format JSON with indentation
        """
        if not rooms:
            return

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to dictionaries
        room_dicts = [room.to_dict() for room in rooms]

        # Build output data
        output_data = {}

        if include_metadata:
            output_data["metadata"] = {
                "export_date": datetime.now().isoformat(),
                "total_rooms": len(rooms),
                "total_area": sum(r.area for r in rooms),
                "total_volume": sum(r.volume for r in rooms),
            }

        output_data["rooms"] = room_dicts

        # Write to file
        with open(output_path, "w") as f:
            if pretty_print:
                json.dump(output_data, f, indent=2, default=str)
            else:
                json.dump(output_data, f, default=str)
